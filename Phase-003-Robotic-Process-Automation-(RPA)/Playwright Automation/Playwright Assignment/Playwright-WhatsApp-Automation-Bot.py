"""Playwright-powered WhatsApp Web message sender and data extractor.

Use this only with contacts who have consented to receive messages.
"""

from __future__ import annotations

import json
import logging
import os
import random
import re
import sys
import time
from collections import Counter
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from playwright.sync_api import (
    BrowserContext,
    Locator,
    Page,
    Playwright,
    TimeoutError as PlaywrightTimeoutError,
    sync_playwright,
)


BASE_DIR = Path(__file__).resolve().parent
CONTACTS_FILE = Path(os.getenv("CONTACTS_FILE", BASE_DIR / "contacts.xlsx"))
PROFILE_DIR = Path(os.getenv("WHATSAPP_PROFILE_DIR", BASE_DIR / ".whatsapp_profile"))
OUTPUT_DIR = Path(os.getenv("OUTPUT_DIR", BASE_DIR / "output"))
SCREENSHOTS_DIR = OUTPUT_DIR / "screenshots"

LOGIN_TIMEOUT_SECONDS = int(os.getenv("LOGIN_TIMEOUT_SECONDS", "300"))
# Required message-flow timing. These are intentionally fixed so environment
# variables from earlier test runs cannot cause premature follow-ups.
REPLY_TIMEOUT_SECONDS = 300
FOLLOW_UP_AFTER_SECONDS = 180
REPLY_POLL_INTERVAL_MS = 500
QUICK_CONFIRMATION_DEADLINE_SECONDS = 10
MAX_INCOMING_BUBBLES_TO_SCAN = 30

DEFAULT_MESSAGE = (
    'Hi {name}! This is a test message from my Playwright WhatsApp '
    'Automation project. Please reply "Hi". Thank you!'
)
EXPECTED_REPLY = "Hi"
FOLLOW_UP_MESSAGE = 'Hi {name}, just checking in. Please reply "Hi" when available.'
CONFIRMATION_MESSAGE = "Thank you {name}! Your reply has been received successfully."

SEARCH_SELECTORS = (
    '[aria-label="Search or start a new chat"]',
    'input[placeholder="Search or start a new chat"]',
    '[role="searchbox"]',
    '#side div[contenteditable="true"][role="textbox"]',
    '#side div[contenteditable="true"]',
    'div[contenteditable="true"][data-tab="3"]',
    'div[aria-label="Search input textbox"]',
)
USE_HERE_SELECTORS = (
    'button:has-text("Use here")',
    '[role="button"]:has-text("Use here")',
)
COMPOSER_SELECTORS = (
    'footer div[contenteditable="true"][role="textbox"]',
    'footer div[contenteditable="true"][data-tab="10"]',
    'div[aria-label="Type a message"]',
)
CHAT_TITLE_SELECTORS = (
    '#main header [data-testid="conversation-info-header-chat-title"]',
    '#main header span[title]',
    '#main header span[dir="auto"]',
)


@dataclass
class Contact:
    name: str
    phone: str
    message_template: str = DEFAULT_MESSAGE


@dataclass
class ContactResult:
    name: str
    phone_masked: str
    status: str = "pending"
    sent_message: str = ""
    follow_up_sent: bool = False
    expected_reply_received: bool = False
    last_three_messages_from_contact: list[dict[str, Any]] = field(default_factory=list)
    screenshot: str = ""
    error: str = ""
    started_at: str = ""
    finished_at: str = ""


def human_pause(page: Page, minimum: float = 2.0, maximum: float = 5.0) -> None:
    """Pause for a small randomized interval between user-like actions."""
    page.wait_for_timeout(int(random.uniform(minimum, maximum) * 1000))


def first_visible(page: Page, selectors: tuple[str, ...], timeout_ms: int) -> Locator:
    """Return the first visible locator among a list of selector fallbacks."""
    deadline = time.monotonic() + timeout_ms / 1000
    while time.monotonic() < deadline:
        for selector in selectors:
            try:
                page.wait_for_selector(selector, state="visible", timeout=500)
                return page.locator(selector).first
            except PlaywrightTimeoutError:
                pass
        page.wait_for_timeout(250)
    raise PlaywrightTimeoutError(f"No visible element found for selectors: {selectors}")


def visible_now(page: Page, selectors: tuple[str, ...]) -> Locator | None:
    """Return a currently visible fallback locator without a long wait."""
    for selector in selectors:
        locator = page.locator(selector).first
        try:
            if locator.is_visible(timeout=250):
                return locator
        except PlaywrightTimeoutError:
            pass
    return None


def normalize_phone(value: Any) -> str:
    """Normalize spreadsheet phone values while retaining a leading plus sign."""
    raw = str(value or "").strip()
    if not raw:
        return ""
    prefix = "+" if raw.startswith("+") else ""
    return prefix + re.sub(r"\D", "", raw)


def mask_phone(phone: str) -> str:
    digits = re.sub(r"\D", "", phone)
    if len(digits) <= 4:
        return "*" * len(digits)
    country_prefix = "+" if phone.startswith("+") else ""
    return f"{country_prefix}{digits[:2]}{'*' * (len(digits) - 4)}{digits[-2:]}"


def message_flow(contact: Contact) -> dict[str, str]:
    """Build every required message from one authoritative flow definition."""
    return {
        "initial": contact.message_template.format(name=contact.name),
        "expected_reply": EXPECTED_REPLY,
        "follow_up": FOLLOW_UP_MESSAGE.format(name=contact.name),
        "confirmation": CONFIRMATION_MESSAGE.format(name=contact.name),
    }


def load_contacts(path: Path) -> list[Contact]:
    if not path.exists():
        raise FileNotFoundError(
            f"Contacts workbook not found: {path}. Copy contacts.example.xlsx to contacts.xlsx first."
        )

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [str(value or "").strip().lower() for value in next(rows)]
    except StopIteration as exc:
        raise ValueError("The contacts workbook is empty.") from exc

    required = {"name", "phone"}
    missing = required.difference(headers)
    if missing:
        raise ValueError(f"Missing required contacts.xlsx column(s): {', '.join(sorted(missing))}")

    name_index = headers.index("name")
    phone_index = headers.index("phone")
    message_index = headers.index("message") if "message" in headers else None
    contacts: list[Contact] = []

    for row_number, row in enumerate(rows, start=2):
        name = str(row[name_index] or "").strip()
        phone = normalize_phone(row[phone_index])
        message = ""
        if message_index is not None and message_index < len(row):
            message = str(row[message_index] or "").strip()
        if not name and not phone:
            continue
        if not name or not phone:
            logging.warning("Skipping incomplete contact on row %s.", row_number)
            continue
        contacts.append(Contact(name=name, phone=phone, message_template=message or DEFAULT_MESSAGE))

    workbook.close()
    if not contacts:
        raise ValueError("No valid contacts were found in the workbook.")
    return contacts


def validate_configuration() -> None:
    if LOGIN_TIMEOUT_SECONDS <= 0:
        raise ValueError("LOGIN_TIMEOUT_SECONDS must be greater than zero.")


def launch_edge(playwright: Playwright) -> BrowserContext:
    PROFILE_DIR.mkdir(parents=True, exist_ok=True)
    return playwright.chromium.launch_persistent_context(
        user_data_dir=str(PROFILE_DIR),
        channel="msedge",
        headless=False,
        chromium_sandbox=True,
        viewport={"width": 1440, "height": 960},
        args=["--start-maximized"],
    )


def wait_for_manual_use_here(page: Page, use_here_button: Locator | None = None) -> bool:
    """Pause for a manual window takeover when WhatsApp requests one."""
    use_here_button = use_here_button or visible_now(page, USE_HERE_SELECTORS)
    if use_here_button is None:
        return False

    logging.warning(
        'WhatsApp is active in another window. In the Edge window, click "Use here" manually. '
        "The automation will continue after your click."
    )
    try:
        use_here_button.wait_for(state="hidden", timeout=LOGIN_TIMEOUT_SECONDS * 1000)
    except PlaywrightTimeoutError as exc:
        raise PlaywrightTimeoutError(
            f'Timed out after {LOGIN_TIMEOUT_SECONDS} seconds waiting for you to click "Use here".'
        ) from exc

    logging.info('"Use here" was clicked. Resuming the automation.')
    page.wait_for_timeout(1_500)
    return True


def wait_for_whatsapp_login(page: Page) -> None:
    page.goto("https://web.whatsapp.com/", wait_until="domcontentloaded", timeout=60_000)
    logging.info(
        'Waiting for WhatsApp Web. Scan the QR code if required, or click "Use here" manually '
        "if WhatsApp asks to move the active session to this window."
    )
    # Let late-loading dialogs render before treating the search control as ready.
    page.wait_for_timeout(2_000)
    deadline = time.monotonic() + LOGIN_TIMEOUT_SECONDS

    while time.monotonic() < deadline:
        use_here_button = visible_now(page, USE_HERE_SELECTORS)
        if use_here_button is not None:
            wait_for_manual_use_here(page, use_here_button)
            continue

        search_box = visible_now(page, SEARCH_SELECTORS)
        if search_box is not None:
            try:
                # A trial click confirms the control is not covered by a modal.
                search_box.click(trial=True, timeout=1_000)
                logging.info("WhatsApp Web is ready. Starting contact processing.")
                return
            except PlaywrightTimeoutError:
                pass

        page.wait_for_timeout(500)

    raise PlaywrightTimeoutError(
        f"WhatsApp Web was not ready after {LOGIN_TIMEOUT_SECONDS} seconds. "
        'Complete the QR login or click "Use here" in the Edge window.'
    )


def digits_only(value: str) -> str:
    return re.sub(r"\D", "", value)


def result_text(locator: Locator) -> str:
    """Read searchable text and title metadata from a result row."""
    try:
        text = locator.inner_text(timeout=1_000).strip()
    except (PlaywrightTimeoutError, AttributeError):
        text = ""
    try:
        title = locator.get_attribute("title", timeout=1_000) or ""
    except (PlaywrightTimeoutError, AttributeError):
        title = ""
    return f"{title} {text}".strip()


def first_chat_search_result(page: Page, phone: str, timeout_ms: int = 20_000) -> Locator:
    """Find the first user row in the Chats section of WhatsApp search results."""
    pane = page.locator("#pane-side")
    target_digits = digits_only(phone)
    deadline = time.monotonic() + timeout_ms / 1000

    while time.monotonic() < deadline:
        # First prefer any titled result that contains the searched number.
        titled_results = pane.locator("span[title]")
        for index in range(min(titled_results.count(), 50)):
            titled = titled_results.nth(index)
            if not titled.is_visible():
                continue
            if target_digits and target_digits in digits_only(result_text(titled)):
                clickable = titled.locator(
                    "xpath=ancestor::*[@role='listitem' or @role='button' or @tabindex][1]"
                )
                return clickable if clickable.count() else titled

        # WhatsApp labels contact matches as "Chats". Select the first row after
        # that heading, supporting both current and older row markup.
        headings = pane.get_by_text("Chats", exact=True)
        for heading_index in range(headings.count()):
            heading = headings.nth(heading_index)
            if not heading.is_visible():
                continue
            following_selectors = (
                "xpath=following::*[@data-testid='cell-frame-container'][1]",
                "xpath=following::*[@role='listitem'][1]",
                "xpath=following::*[@role='button' and (@tabindex='0' or @tabindex='-1')][1]",
                "xpath=following::*[@tabindex='-1'][1]",
                "xpath=following::span[@title][1]",
            )
            for selector in following_selectors:
                candidate = heading.locator(selector).first
                try:
                    if candidate.is_visible(timeout=250):
                        return candidate
                except PlaywrightTimeoutError:
                    pass

        # Fallbacks for layouts that omit the Chats heading. Search only inside
        # the filtered results pane and prefer a row containing the phone digits.
        rows = pane.locator(
            "[data-testid='cell-frame-container'], div[role='listitem'], "
            "div[role='button'][tabindex], div[tabindex='-1']"
        )
        fallback: Locator | None = None
        for index in range(min(rows.count(), 50)):
            row = rows.nth(index)
            try:
                if not row.is_visible(timeout=250):
                    continue
            except PlaywrightTimeoutError:
                continue
            text = result_text(row)
            if not text:
                continue
            if fallback is None:
                fallback = row
            if target_digits and target_digits in digits_only(text):
                return row
        if fallback is not None:
            return fallback

        page.wait_for_timeout(500)

    raise LookupError(f"No result appeared in the WhatsApp Chats section for {mask_phone(phone)}.")


def open_chat_name(page: Page, fallback_name: str) -> str:
    """Read the displayed contact name from the opened conversation header."""
    deadline = time.monotonic() + 10
    ignored = {"online", "offline", "typing…", "typing...", "click here for contact info"}
    while time.monotonic() < deadline:
        for selector in CHAT_TITLE_SELECTORS:
            titles = page.locator(selector)
            for index in range(min(titles.count(), 10)):
                title = titles.nth(index)
                try:
                    if not title.is_visible(timeout=250):
                        continue
                    value = (title.get_attribute("title") or title.inner_text()).strip()
                except PlaywrightTimeoutError:
                    continue
                if value and value.casefold() not in ignored:
                    return value
        page.wait_for_timeout(250)
    logging.warning("Could not read the chat header name; using the workbook name instead.")
    return fallback_name


def search_and_open_chat(page: Page, contact: Contact) -> str:
    logging.info("Opening the chat search for %s.", contact.name)
    search_box = first_visible(page, SEARCH_SELECTORS, 20_000)
    search_box.click()
    search_box.press("Control+A")
    search_box.press("Backspace")
    human_pause(page)
    search_box.type(contact.phone, delay=random.randint(55, 110))
    logging.info("Search entered; waiting for a matching chat.")

    selected = first_chat_search_result(page, contact.phone)
    selected.click()
    logging.info("First Chats result opened; waiting for the message composer.")
    first_visible(page, COMPOSER_SELECTORS, 20_000)
    resolved_name = open_chat_name(page, contact.name)
    logging.info("Chat recipient identified as %s.", resolved_name)
    human_pause(page)
    return resolved_name


def outgoing_signature(messages: Locator) -> tuple[str, str] | None:
    """Return a stable-enough signature for the latest outgoing bubble."""
    if messages.count() == 0:
        return None
    latest = messages.last
    try:
        message_id = latest.get_attribute("data-id") or ""
        if not message_id:
            id_nodes = latest.locator("[data-id]")
            if id_nodes.count():
                message_id = id_nodes.first.get_attribute("data-id") or ""
        text = latest.inner_text(timeout=1_000).strip()
        return message_id, text
    except PlaywrightTimeoutError:
        return None


def composer_is_empty(composer: Locator) -> bool:
    """Check whether WhatsApp cleared the message composer after Enter."""
    try:
        content = composer.text_content(timeout=500) or ""
    except (PlaywrightTimeoutError, AttributeError):
        return False
    content = re.sub(r"[\s\u200b-\u200d\ufeff]", "", content)
    return content == ""


def send_message(page: Page, message: str, *, quick: bool = False) -> None:
    composer = first_visible(page, COMPOSER_SELECTORS, 15_000)
    # The current WhatsApp layout no longer consistently provides #main.
    outgoing = page.locator("div.message-out")
    before_signature = outgoing_signature(outgoing)

    composer.click()
    # fill() inserts Unicode names reliably; keyboard typing can silently drop
    # characters that are unavailable on the active keyboard layout.
    composer.fill(message)
    page.wait_for_timeout(100 if quick else random.randint(500, 1_200))
    composer.press("Enter")

    # Confirm a new outgoing bubble instead of matching the entire message text.
    # WhatsApp can split or normalize Unicode text across nested spans.
    confirmation_timeout = QUICK_CONFIRMATION_DEADLINE_SECONDS if quick else 20
    deadline = time.monotonic() + confirmation_timeout
    composer_cleared_at: float | None = None
    while time.monotonic() < deadline:
        after_signature = outgoing_signature(outgoing)
        if after_signature is not None and after_signature != before_signature:
            logging.info("WhatsApp confirmed a new outgoing message bubble.")
            if not quick:
                human_pause(page)
            return
        if composer_is_empty(composer):
            composer_cleared_at = composer_cleared_at or time.monotonic()
            if time.monotonic() - composer_cleared_at >= 1:
                logging.info(
                    "WhatsApp cleared the composer after Enter; treating the message as sent."
                )
                if not quick:
                    human_pause(page)
                return
        else:
            composer_cleared_at = None
        page.wait_for_timeout(250)

    raise PlaywrightTimeoutError(
        "The message was entered, but WhatsApp did not confirm it within "
        f"{confirmation_timeout} seconds."
    )


def bubble_text_candidates(bubble: Locator) -> list[str]:
    """Collect possible message-text segments from current and older layouts."""
    text_selectors = (
        "div.copyable-text[data-pre-plain-text] span.selectable-text",
        "div.copyable-text[data-pre-plain-text] span[dir='ltr']",
        "div.copyable-text[data-pre-plain-text] span[dir='auto']",
        "[data-testid='msg-text'] span.selectable-text",
        "[data-testid='msg-text']",
        "span.selectable-text",
    )
    candidates: list[str] = []
    for selector in text_selectors:
        for part in bubble.locator(selector).all_inner_texts():
            value = part.strip()
            if value and value not in candidates:
                candidates.append(value)
            for line in value.splitlines():
                line = line.strip()
                if line and line not in candidates:
                    candidates.append(line)
    try:
        for line in bubble.inner_text(timeout=1_000).splitlines():
            line = line.strip()
            if line and line not in candidates:
                candidates.append(line)
    except (PlaywrightTimeoutError, AttributeError):
        pass
    return candidates


def actual_bubble_text(bubble: Locator) -> str:
    """Extract the reply itself, excluding quoted-message text when possible."""
    preferred = bubble.locator(
        "div.copyable-text[data-pre-plain-text] span.selectable-text, "
        "[data-testid='msg-text'] span.selectable-text, span.selectable-text"
    ).all_inner_texts()
    preferred = [part.strip() for part in preferred if part.strip()]
    if preferred:
        return preferred[-1]
    candidates = bubble_text_candidates(bubble)
    return candidates[0] if candidates else ""


def incoming_messages(page: Page) -> list[dict[str, Any]]:
    messages: list[dict[str, Any]] = []
    # Do not scope to #main: newer WhatsApp layouts omit that container ID.
    bubbles = page.locator("div.message-in")
    bubble_count = bubbles.count()
    first_index = max(0, bubble_count - MAX_INCOMING_BUBBLES_TO_SCAN)
    for index in range(first_index, bubble_count):
        bubble = bubbles.nth(index)
        try:
            candidates = bubble_text_candidates(bubble)
            text = actual_bubble_text(bubble)
            copyable = bubble.locator("div.copyable-text[data-pre-plain-text]")
            metadata = (
                copyable.first.get_attribute("data-pre-plain-text") or "" if copyable.count() else ""
            )
            message_id = bubble.get_attribute("data-id") or ""
            if not message_id:
                id_nodes = bubble.locator("[data-id]")
                if id_nodes.count():
                    message_id = id_nodes.first.get_attribute("data-id") or ""
            if text:
                messages.append(
                    {"text": text, "candidates": candidates, "metadata": metadata, "id": message_id}
                )
        except Exception as exc:  # A bubble can disappear while the chat virtualizes.
            logging.debug("Skipped a transient message bubble: %s", exc)
    return messages


def is_expected_reply(text: str) -> bool:
    normalized = re.sub(r"[\u200b-\u200d\ufeff]", "", text).strip().lower()
    return normalized.rstrip("!.,").strip() == EXPECTED_REPLY.lower()


def message_fingerprint(message: dict[str, Any]) -> tuple[str, str, str]:
    return (message.get("id", ""), message.get("metadata", ""), message.get("text", ""))


def wait_for_reply(
    page: Page, contact: Contact, incoming_before: list[dict[str, Any]]
) -> tuple[bool, bool]:
    started = time.monotonic()
    follow_up_sent = False
    flow = message_flow(contact)
    baseline_counts = Counter(message_fingerprint(message) for message in incoming_before)
    observed_new: set[tuple[str, str, str]] = set()
    next_progress_log = 10
    logging.info(
        'Initial message sent. Waiting up to %s seconds for a new "Hi" reply; follow-up is due after %s seconds.',
        REPLY_TIMEOUT_SECONDS,
        FOLLOW_UP_AFTER_SECONDS,
    )

    while time.monotonic() - started < REPLY_TIMEOUT_SECONDS:
        current = incoming_messages(page)
        current_counts: Counter[tuple[str, str, str]] = Counter()
        for message in current:
            fingerprint = message_fingerprint(message)
            current_counts[fingerprint] += 1
            if current_counts[fingerprint] <= baseline_counts[fingerprint]:
                continue

            candidates = message.get("candidates") or [message["text"]]
            if any(is_expected_reply(str(candidate)) for candidate in candidates):
                logging.info('New "Hi" reply received from %s.', contact.name)
                return True, follow_up_sent
            if fingerprint not in observed_new:
                logging.info('A new incoming message was detected, but it was not an exact "Hi" reply.')
                observed_new.add(fingerprint)

        elapsed = time.monotonic() - started
        elapsed_seconds = int(elapsed)
        if elapsed_seconds >= next_progress_log:
            remaining = max(0, REPLY_TIMEOUT_SECONDS - elapsed_seconds)
            logging.info(
                'Waiting for "Hi": %s seconds elapsed, %s seconds remaining.',
                elapsed_seconds,
                remaining,
            )
            next_progress_log = elapsed_seconds + 10
        if not follow_up_sent and elapsed >= FOLLOW_UP_AFTER_SECONDS:
            logging.info("MESSAGE FLOW 3/4 — sending follow-up: %s", flow["follow_up"])
            send_message(page, flow["follow_up"])
            follow_up_sent = True
            logging.info("Follow-up sent to %s.", contact.name)

        # Polling is read-only and can be frequent; user-like delays are retained
        # for browser actions, while reply detection stays responsive.
        page.wait_for_timeout(REPLY_POLL_INTERVAL_MS)

    return False, follow_up_sent


def safe_filename(value: str) -> str:
    clean = re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("._")
    return clean or "contact"


def capture_screenshot(page: Page, contact: Contact) -> str:
    SCREENSHOTS_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{safe_filename(contact.name)}_{timestamp}.png"
    path = SCREENSHOTS_DIR / filename
    page.screenshot(path=str(path), full_page=False)
    return str(path.relative_to(BASE_DIR))


def process_contact(page: Page, contact: Contact) -> ContactResult:
    result = ContactResult(
        name=contact.name,
        phone_masked=mask_phone(contact.phone),
        started_at=datetime.now().isoformat(timespec="seconds"),
    )
    try:
        logging.info("Processing %s (%s).", contact.name, result.phone_masked)
        resolved_name = search_and_open_chat(page, contact)
        contact.name = resolved_name
        result.name = resolved_name
        incoming_before = incoming_messages(page)
        flow = message_flow(contact)
        result.sent_message = flow["initial"]
        logging.info("MESSAGE FLOW 1/4 — sending initial message: %s", flow["initial"])
        send_message(page, result.sent_message)

        logging.info('MESSAGE FLOW 2/4 — expected user reply: "%s".', flow["expected_reply"])
        reply_received, follow_up_sent = wait_for_reply(page, contact, incoming_before)
        result.follow_up_sent = follow_up_sent
        result.expected_reply_received = reply_received
        if reply_received:
            logging.info("MESSAGE FLOW 4/4 — sending quick confirmation: %s", flow["confirmation"])
            confirmation_started = time.monotonic()
            send_message(page, flow["confirmation"], quick=True)
            logging.info(
                "Successful reply confirmation completed %.2f seconds after detection.",
                time.monotonic() - confirmation_started,
            )
            result.status = "reply_received"
        else:
            logging.info('No new "Hi" reply arrived before the reply window ended.')
            result.status = "no_reply"

        result.last_three_messages_from_contact = incoming_messages(page)[-3:]
        logging.info("Taking the final conversation screenshot for %s.", contact.name)
        result.screenshot = capture_screenshot(page, contact)
    except Exception as exc:
        result.status = "failed"
        result.error = str(exc)
        logging.exception("Could not process %s.", contact.name)
        try:
            result.screenshot = capture_screenshot(page, contact)
        except Exception:
            logging.exception("Could not capture the failure screenshot.")
    finally:
        result.finished_at = datetime.now().isoformat(timespec="seconds")
    return result


def unlocked_report_fallback(path: Path) -> Path:
    """Build a collision-resistant report name when the dated file is open."""
    timestamp = datetime.now().strftime("%H%M%S_%f")
    return path.with_name(f"{path.stem}_{timestamp}{path.suffix}")


def save_reports(results: list[ContactResult]) -> tuple[Path, Path]:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    report_date = date.today().isoformat()
    json_path = OUTPUT_DIR / f"whatsapp_report_{report_date}.json"
    excel_path = OUTPUT_DIR / f"whatsapp_report_{report_date}.xlsx"

    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "total_contacts": len(results),
        "summary": {
            "successful": sum(item.status != "failed" for item in results),
            "failed": sum(item.status == "failed" for item in results),
            "replies_received": sum(item.expected_reply_received for item in results),
        },
        "contacts": [asdict(item) for item in results],
    }
    json_content = json.dumps(payload, indent=2, ensure_ascii=False)
    try:
        json_path.write_text(json_content, encoding="utf-8")
    except PermissionError:
        locked_path = json_path
        json_path = unlocked_report_fallback(json_path)
        logging.warning("%s is locked; saving this run to %s instead.", locked_path.name, json_path.name)
        json_path.write_text(json_content, encoding="utf-8")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "WhatsApp Report"
    sheet.append(
        [
            "Name",
            "Phone (masked)",
            "Status",
            "Follow-up sent",
            "Reply received",
            "Last 3 incoming messages",
            "Screenshot",
            "Error",
            "Started at",
            "Finished at",
        ]
    )
    for item in results:
        last_messages = " | ".join(message["text"] for message in item.last_three_messages_from_contact)
        sheet.append(
            [
                item.name,
                item.phone_masked,
                item.status,
                item.follow_up_sent,
                item.expected_reply_received,
                last_messages,
                item.screenshot,
                item.error,
                item.started_at,
                item.finished_at,
            ]
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = {"A": 22, "B": 20, "C": 20, "D": 16, "E": 16, "F": 55, "G": 42, "H": 45, "I": 22, "J": 22}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    try:
        workbook.save(excel_path)
    except PermissionError:
        locked_path = excel_path
        excel_path = unlocked_report_fallback(excel_path)
        logging.warning("%s is locked; saving this run to %s instead.", locked_path.name, excel_path.name)
        workbook.save(excel_path)
    finally:
        workbook.close()
    return json_path, excel_path


def main() -> int:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
    context: BrowserContext | None = None
    results: list[ContactResult] = []

    try:
        validate_configuration()
        contacts = load_contacts(CONTACTS_FILE)
        logging.info("Loaded %s contact(s).", len(contacts))
        with sync_playwright() as playwright:
            context = launch_edge(playwright)
            page = context.pages[0] if context.pages else context.new_page()
            wait_for_whatsapp_login(page)
            for contact in contacts:
                results.append(process_contact(page, contact))
            context.close()
            context = None
    except KeyboardInterrupt:
        logging.warning("Run cancelled by the user. Saving completed results.")
    except Exception as exc:
        logging.exception("Automation could not start or complete: %s", exc)
    finally:
        if context is not None:
            try:
                context.close()
            except Exception:
                logging.debug("Browser context was already closed.", exc_info=True)

    if results:
        json_path, excel_path = save_reports(results)
        logging.info("Reports saved to %s and %s.", json_path, excel_path)
        return 0 if all(item.status != "failed" for item in results) else 1
    logging.error("No contact results were produced.")
    return 1


if __name__ == "__main__":
    sys.exit(main())
