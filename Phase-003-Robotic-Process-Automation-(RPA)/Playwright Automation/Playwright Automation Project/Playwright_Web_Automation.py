"""SauceDemo product extraction, screenshot, and Excel automation.

Run visibly (the project requirement and the default):
    python Playwright_Web_Automation.py

Run without displaying the browser:
    python Playwright_Web_Automation.py --headless
"""

from __future__ import annotations

import argparse
import logging
import os
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.sync_api import Browser, Page, Playwright, sync_playwright


BASE_URL = "https://www.saucedemo.com/"
DEFAULT_USERNAME = "standard_user"
DEFAULT_PASSWORD = "secret_sauce"
HEADERS = (
    "Product Name",
    "Price",
    "Product URL",
    "Date",
    "Time",
    "Screenshot File",
    "Status",
)
WINDOWS_RESERVED_NAMES = {
    "CON",
    "PRN",
    "AUX",
    "NUL",
    *(f"COM{number}" for number in range(1, 10)),
    *(f"LPT{number}" for number in range(1, 10)),
}


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract every SauceDemo product and create screenshots and Excel output."
    )
    parser.add_argument(
        "--headless",
        action="store_true",
        help="Run Chromium without displaying its window (headed mode is the default).",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(__file__).resolve().parent / "products.xlsx",
        help="Excel output path (default: products.xlsx beside this script).",
    )
    parser.add_argument(
        "--screenshots",
        type=Path,
        default=Path(__file__).resolve().parent / "screenshots",
        help="Screenshot directory (default: screenshots beside this script).",
    )
    return parser.parse_args()


def sanitize_filename(product_name: str, max_stem_length: int = 120) -> str:
    """Return a non-empty Windows-safe PNG filename based on a product name."""
    stem = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", product_name).rstrip(" .")
    stem = stem[:max_stem_length].rstrip(" .")

    if not stem:
        stem = "product"
    if stem.upper() in WINDOWS_RESERVED_NAMES:
        stem = f"_{stem}"

    return f"{stem}.png"


def unique_screenshot_path(
    directory: Path, product_name: str, claimed_names: set[str]
) -> Path:
    """Choose a screenshot path without overwriting files from this or an older run."""
    safe_name = sanitize_filename(product_name)
    stem = Path(safe_name).stem
    candidate = directory / safe_name
    sequence = 2

    while candidate.name.casefold() in claimed_names or candidate.exists():
        suffix = f"_{sequence}"
        shortened_stem = stem[: 120 - len(suffix)].rstrip(" .") or "product"
        candidate = directory / f"{shortened_stem}{suffix}.png"
        sequence += 1

    claimed_names.add(candidate.name.casefold())
    return candidate


def create_workbook() -> tuple[Workbook, object]:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Products"
    worksheet.append(HEADERS)

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    for cell in worksheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = "A1:G1"
    return workbook, worksheet


def append_result(worksheet: object, result: dict[str, str]) -> None:
    worksheet.append([result.get(header, "") for header in HEADERS])
    url_cell = worksheet.cell(row=worksheet.max_row, column=3)
    if url_cell.value:
        url_cell.hyperlink = str(url_cell.value)
        url_cell.style = "Hyperlink"


def save_workbook(workbook: Workbook, worksheet: object, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    widths = (32, 12, 62, 14, 16, 38, 55)
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    workbook.save(output_path)


def add_url_overlay(page: Page, product_url: str) -> None:
    """Put the detail URL inside the page because Playwright omits the address bar."""
    page.evaluate(
        """
        (url) => {
            document.getElementById('automation-url-overlay')?.remove();
            const overlay = document.createElement('div');
            overlay.id = 'automation-url-overlay';
            overlay.textContent = `Product URL: ${url}`;
            Object.assign(overlay.style, {
                position: 'fixed', top: '0', left: '0', right: '0',
                zIndex: '2147483647', padding: '10px 14px',
                color: '#ffffff', background: '#1e3a5f',
                font: '14px Arial, sans-serif', overflowWrap: 'anywhere',
                pointerEvents: 'none',
                boxShadow: '0 2px 5px rgba(0,0,0,.35)'
            });
            document.body.appendChild(overlay);
        }
        """,
        product_url,
    )


def login(page: Page, username: str, password: str) -> None:
    page.goto(BASE_URL, wait_until="domcontentloaded")
    page.get_by_test_id("username").fill(username)
    page.get_by_test_id("password").fill(password)
    page.get_by_test_id("login-button").click()
    page.wait_for_url("**/inventory.html")
    page.locator(".inventory_list").wait_for(state="visible")


def collect_product_targets(page: Page) -> list[dict[str, str]]:
    """Collect stable link identifiers before visiting the product pages."""
    cards = page.locator(".inventory_item")
    cards.first.wait_for(state="visible")
    targets: list[dict[str, str]] = []

    for index in range(cards.count()):
        card = cards.nth(index)
        name = f"Unknown Product {index + 1}"
        link_test_id = ""
        try:
            extracted_name = (
                card.get_by_test_id("inventory-item-name").inner_text().strip()
            )
            if extracted_name:
                name = extracted_name
        except Exception as error:
            logging.warning("Catalog product %d has no readable name: %s", index + 1, error)

        try:
            link_locator = card.locator("a[data-test$='-title-link']").first
            link_test_id = link_locator.get_attribute("data-test") or ""
        except Exception as error:
            logging.warning("Catalog product %d has no detail link: %s", index + 1, error)

        targets.append(
            {
                "listing_name": name,
                "link_test_id": link_test_id,
            }
        )

    return targets


def process_product(
    page: Page,
    target: dict[str, str],
    screenshot_directory: Path,
    claimed_names: set[str],
) -> dict[str, str]:
    result = {header: "" for header in HEADERS}
    result["Product Name"] = target.get("listing_name", "Unknown Product")
    issues: list[str] = []

    processed_at = datetime.now()
    result["Date"] = processed_at.strftime("%d-%m-%Y")
    result["Time"] = processed_at.strftime("%I:%M:%S %p")

    if not target.get("link_test_id"):
        issues.append("Product detail link not found")
        result["Status"] = "Failed - " + "; ".join(issues)
        return result

    try:
        # SauceDemo's catalog anchors use href="#" and perform routing in their
        # click handlers, so clicking is required to obtain the real detail URL.
        page.get_by_test_id(target["link_test_id"]).click()
        page.wait_for_url("**/inventory-item.html?id=*")
        page.get_by_test_id("inventory-item-name").wait_for(state="visible")
        result["Product URL"] = page.url
    except Exception as error:
        issues.append(f"Product page navigation error: {error}")
        result["Status"] = "Failed - " + "; ".join(issues)
        return result

    try:
        detail_name = page.get_by_test_id("inventory-item-name").inner_text().strip()
        if detail_name:
            result["Product Name"] = detail_name
        else:
            issues.append("Product name not found")
    except Exception as error:
        issues.append(f"Product name not found: {error}")

    try:
        result["Price"] = (
            page.get_by_test_id("inventory-item-price").inner_text().strip()
        )
        if not result["Price"]:
            issues.append("Product price not found")
    except Exception as error:
        issues.append(f"Product price not found: {error}")

    screenshot_path = unique_screenshot_path(
        screenshot_directory, result["Product Name"], claimed_names
    )
    result["Screenshot File"] = screenshot_path.name
    try:
        add_url_overlay(page, result["Product URL"])
        page.screenshot(path=str(screenshot_path), full_page=True)
    except Exception as error:
        issues.append(f"Screenshot could not be saved: {error}")

    result["Status"] = "Success" if not issues else "Failed - " + "; ".join(issues)
    return result


def run_automation(
    playwright: Playwright,
    worksheet: object,
    screenshot_directory: Path,
    headless: bool,
) -> None:
    browser: Browser | None = None
    try:
        # SauceDemo exposes stable `data-test` attributes rather than
        # Playwright's default `data-testid` attribute.
        playwright.selectors.set_test_id_attribute("data-test")
        # browser = playwright.chromium.launch(headless=headless)
        browser = playwright.chromium.launch(headless=False)
        context = browser.new_context(viewport={"width": 1440, "height": 1000})
        page = context.new_page()
        page.set_default_timeout(15_000)
        page.set_default_navigation_timeout(30_000)

        username = os.getenv("SAUCEDEMO_USERNAME", DEFAULT_USERNAME)
        password = os.getenv("SAUCEDEMO_PASSWORD", DEFAULT_PASSWORD)
        login(page, username, password)
        targets = collect_product_targets(page)
        if not targets:
            raise RuntimeError("No products were found on the inventory page")

        logging.info("Found %d products", len(targets))
        claimed_names: set[str] = set()
        for number, target in enumerate(targets, start=1):
            try:
                result = process_product(
                    page, target, screenshot_directory, claimed_names
                )
            except Exception as error:
                now = datetime.now()
                result = {
                    "Product Name": target.get("listing_name", "Unknown Product"),
                    "Price": "",
                    "Product URL": "",
                    "Date": now.strftime("%d-%m-%Y"),
                    "Time": now.strftime("%I:%M:%S %p"),
                    "Screenshot File": "",
                    "Status": f"Failed - Unexpected product error: {error}",
                }
            append_result(worksheet, result)
            logging.info(
                "[%d/%d] %s: %s",
                number,
                len(targets),
                result["Product Name"],
                result["Status"],
            )
            # Return through the UI so the next product link can be clicked. If a
            # failed detail page has no back button, reload the authenticated list.
            try:
                if "/inventory-item.html" in page.url:
                    page.get_by_test_id("back-to-products").click()
                    page.wait_for_url("**/inventory.html")
                elif "/inventory.html" not in page.url:
                    page.goto(f"{BASE_URL}inventory.html")
                page.locator(".inventory_list").wait_for(state="visible")
            except Exception as error:
                logging.warning("Could not restore product listing: %s", error)
    finally:
        if browser is not None:
            browser.close()


def main() -> int:
    args = parse_arguments()
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    args.screenshots.mkdir(parents=True, exist_ok=True)
    workbook, worksheet = create_workbook()
    automation_error: Exception | None = None

    try:
        with sync_playwright() as playwright:
            run_automation(playwright, worksheet, args.screenshots, args.headless)
    except Exception as error:
        automation_error = error
        now = datetime.now()
        append_result(
            worksheet,
            {
                "Product Name": "Automation Error",
                "Date": now.strftime("%d-%m-%Y"),
                "Time": now.strftime("%I:%M:%S %p"),
                "Status": f"Failed - {error}",
            },
        )
        logging.exception("Automation stopped before all products could be processed")

    try:
        save_workbook(workbook, worksheet, args.output)
        logging.info("Excel report saved to %s", args.output.resolve())
    except Exception as error:
        logging.error("Excel report could not be saved: %s", error)
        return 1

    return 1 if automation_error else 0


if __name__ == "__main__":
    raise SystemExit(main())
